import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.plotarea import DataTable

def create_excel(year, language_code, texts, with_dashboard=False):
    
    if with_dashboard:
        FILENAME = f"{texts['filename']}.xlsx"
    else:
        FILENAME = f"{texts['filename']}_{year}.xlsx"

    wb = openpyxl.Workbook()
    
    # Styles Setup (c = colors)
    c = {
        "navy": "FF2C3E50", "blue": "FF3498DB", "invest": "FF8E44AD",
        "green": "FF27AE60", "orange": "FFE67E22", "red": "FFC0392B",
        "white": "FFFFFFFF", "gray": "FF7F8C8D", "border": "FFD1D5DB"
    }
    border = Border(left=Side(style='thin', color=c["border"]), right=Side(style='thin', color=c["border"]), 
                    top=Side(style='thin', color=c["border"]), bottom=Side(style='thin', color=c["border"]))
    FMT_CURRENCY = '#,##0.00 "€"' 
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

    # --- SHEET: MAIN PLANNER ---
    ws_main = wb.active
    ws_main.title = str(year)
    ws_main.sheet_view.showGridLines = False
    ws_main.sheet_view.showRowColHeaders = False

    # Header
    ws_main["B2"] = f"{texts['header_title']} {year}"
    ws_main["B2"].font = Font(size=22, bold=True, color=c["navy"])

    # 1. Assets top left
    ws_main["B4"] = texts['assets_title']
    ws_main["B4"].font = Font(bold=True)
    
    col_map = {0: "AF", 1: "AG", 2: "AH", 3: "AI"} 
    start_vals = [3500.50, 12000.00, 500.00, 45000.00, 5500.00] 
    
    for i, label in enumerate(texts['assets_labels']):
        r = 5 + i
        ws_main.cell(row=r, column=2, value=label)
        c_val = ws_main.cell(row=r, column=3)
        c_val.font = Font(bold=True); c_val.number_format = FMT_CURRENCY
        
        if i == 4: # Net Income - fixed start value
            c_val.value = start_vals[i]
        else: # Dynamically pull from Log
            col = col_map[i]
            c_val.value = f"=INDEX(${col}$21:${col}$32, MONTH(TODAY()))"

    # 2. KPIs
    def create_kpi(cell, label, formula, color, fmt=FMT_CURRENCY):
        ws_main[cell] = label; ws_main[cell].font = Font(size=9, color=c["gray"])
        v = ws_main.cell(row=ws_main[cell].row+1, column=ws_main[cell].column, value=formula)
        v.font = Font(size=14, bold=True, color=color); v.number_format = fmt

    # Fixed costs real formula 
    fix_real_formula = "=(SUM(F21:F100) + SUM(M21:M100) + SUM(T21:T100) + SUM(AA21:AA100))"
    # Buffer = G(Living) + N(Digital) + U(Insurance) + AB(Invest Buffer)
    buffer_formula = "=(SUM(G21:G100) + SUM(N21:N100) + SUM(U21:U100) + SUM(AB21:AB100))"
    
    # NEW: Catch division by zero for runway (IFERROR + IF)
    # If E5 (Fixed costs) is 0, show "∞", else calculate normally.
    runway_formula = '=IF(E5=0, "∞", (C5+C6+C7)/E5)'

    create_kpi("E4", texts['kpi_fix_real'], fix_real_formula, c["red"])
    create_kpi("G4", texts['kpi_buffer'], buffer_formula, c["orange"])
    create_kpi("I4", texts['kpi_runway'], runway_formula, c["green"], f'0.0 "{texts["unit_months"]}"')
    create_kpi("K4", texts['kpi_save_rate'], "=(SUM(Y21:Y100)+(C9-(E5+SUM(Y21:Y100))))/C9", c["blue"], "0.0%")
    create_kpi("M4", texts['kpi_total'], "=SUM(C5:C8)", c["navy"])

    # 3. Cashflow Preview
    ws_main["B12"] = texts['preview_title']
    ws_main["B12"].font = Font(bold=True)
    months = texts['months']
    
    for i, m in enumerate(months):
        col = 2 + i
        cell = ws_main.cell(row=13, column=col, value=m)
        cell.fill = PatternFill("solid", c["navy"]); cell.font = Font(color=c["white"], bold=True); cell.alignment = ALIGN_CENTER
        
        wild = f"\"*{m}*\""
        f_all = f"=(SUMPRODUCT((C21:C100=1)*(D21:D100)) + SUMIF(E21:E100,{wild},D21:D100)) + " \
                f"(SUMPRODUCT((J21:J100=1)*(K21:K100)) + SUMIF(L21:L100,{wild},K21:K100)) + " \
                f"(SUMPRODUCT((Q21:Q100=1)*(R21:R100)) + SUMIF(S21:S100,{wild},R21:R100)) + " \
                f"(SUMPRODUCT((X21:X100=1)*(Y21:Y100)) + SUMIF(Z21:Z100,{wild},Y21:Y100))"
        ws_main.cell(row=14, column=col, value=f_all).number_format = FMT_CURRENCY
        ws_main.cell(row=14, column=col).alignment = ALIGN_CENTER

    # 4. blocks for categories
    def draw_block(col, row, title, headers, data, color, is_invest=False, is_log=False):
        ws_main.cell(row=row-1, column=col, value=title).font = Font(bold=True)
        for i, h in enumerate(headers):
            cell = ws_main.cell(row=row, column=col+i, value=h)
            cell.fill = PatternFill("solid", color); cell.font = Font(color=c["white"], bold=True); cell.alignment = ALIGN_CENTER
        for r in range(20):
            curr_r = row+1+r
            for c_idx in range(len(headers)):
                cell = ws_main.cell(row=curr_r, column=col+c_idx); cell.border = border
                if r < len(data) and c_idx < len(data[r]): cell.value = data[r][c_idx]
                
                if not is_log:
                    # Format
                    if is_invest:
                         if c_idx == 2 or c_idx >= 4: cell.number_format = FMT_CURRENCY
                         if c_idx in [1,3]: cell.alignment = ALIGN_CENTER
                    else:
                        # Standard
                        if c_idx in [2,4,5]: cell.number_format = FMT_CURRENCY
                        if c_idx in [1,3]: cell.alignment = ALIGN_CENTER
            
            if not is_log:
                ws_main.cell(row=curr_r, column=col+4, value=f"=IF({get_column_letter(col+1)}{curr_r}>0,{get_column_letter(col+2)}{curr_r}/{get_column_letter(col+1)}{curr_r},0)").number_format = FMT_CURRENCY
                ws_main.cell(row=curr_r, column=col+5, value=f"={get_column_letter(col+4)}{curr_r}*1.05").number_format = FMT_CURRENCY

    # Prepare data
    h_std = [texts['col_item'], texts['col_freq'], texts['col_amount'], texts['col_due'], texts['col_exact'], texts['col_buffer']]
    
    # Header for Invest block also extended with Exact & Buffer
    h_invest = [texts['col_class'], texts['col_freq'], texts['col_amount'], texts['col_due'], texts['col_exact'], texts['col_buffer'], texts['col_goal']]

    due_gez = f"{months[1]}, {months[4]}, {months[7]}, {months[10]}" 
    
    draw_block(2, 20, texts['cat_living'], h_std, [[texts['sample_rent'],1,1450.00,months[0]],[texts['sample_gez'],3,55.00,due_gez]], c["navy"])
    draw_block(9, 20, texts['cat_digital'], h_std, [[texts['sample_netflix'],1,17.99,months[0]],[texts['sample_hosting'],12,120.00,months[1]]], c["blue"])
    draw_block(16, 20, texts['cat_insurance'], h_std, [[texts['sample_kfz'],12,180.00,months[4]]], "FF16A085")
    
    # INVEST BLOCK (Now with calculation columns)
    # Starts at column 23 (W).
    # Data needs to be adjusted: Asset, Frequency, Amount, Due ... Exact/Buffer filled by Excel ... Goal
    draw_block(23, 20, texts['cat_invest'], h_invest, [[texts['sample_etf'],1,1000.00,months[0], None, None, "Rente"]], c["invest"], True)

    # Log Block
    log_data = []
    for i, m in enumerate(months):
        if i == 0: row = [m, start_vals[0], start_vals[1], start_vals[2], start_vals[3], 0, 0]
        else: row = [m, None, None, None, None, None, None]
        log_data.append(row)
    
    # Log Block now starts further to the right because Invest has become wider
    # Invest had 5 columns (W,X,Y,Z,AA). Now 7 columns (W,X,Y,Z,AA,AB,AC).
    # That means Log should start at column AD (30), not 29.
    # We'll leave it at 31 (AD+1) for some breathing room, or directly AD (30).
    draw_block(31, 20, texts['cat_log'], texts['log_cols'], log_data, "FF34495E", False, True)

    # Adjust column widths 
    for i in range(1,50): ws_main.column_dimensions[get_column_letter(i)].width = 13
    
    # Width for text columns (Item/Asset)
    # Living(2), Digital(9), Insurance(16), Invest(23)
    for i in [2,9,16,23]: ws_main.column_dimensions[get_column_letter(i)].width = 25
    
    # Width for KPIs
    for i in [7,14,21,13,5,9,11]: ws_main.column_dimensions[get_column_letter(i)].width = 21

    # --- OPTIONAL: DASHBOARD ---
    if with_dashboard:
        ws_dash = wb.create_sheet("START", 0)
        ws_dash.sheet_view.showGridLines = False
        ws_dash.sheet_view.showRowColHeaders = False
        
        for r in range(1, 4):
            for cx in range(1, 200): ws_dash.cell(row=r, column=cx).fill = PatternFill("solid", c["navy"])
        ws_dash["B2"] = texts['dash_title']; ws_dash["B2"].font = Font(size=24, bold=True, color=c["white"])

        btns = [[f"{year} {texts['btn_open']}", str(year)], [f"{int(year)+1} (Platzhalter)", str(year)]]
        
        for i, b in enumerate(btns):
            cell = ws_dash.cell(row=6+(i*3), column=2, value=f"📂 {b[0]}")
            cell.font = Font(bold=True, color=c["white"], size=12)
            cell.fill = PatternFill("solid", c["blue"])
            cell.alignment = ALIGN_CENTER
            cell.hyperlink = f"#'{b[1]}'!A1"
            ws_dash.cell(row=7+(i*3), column=2).fill = PatternFill("solid", c["gray"])

        ws_dash.column_dimensions["B"].width = 30; ws_dash.column_dimensions["C"].width = 5

        # Backend Data
        # Log Columns: AE=Month, AF=Checking, AG=Daily, AH=Cash, AI=Depot
        ws_dash["X5"] = "Month"; ws_dash["Y5"] = "Total"
        for i in range(12):
            ws_dash[f"X{6+i}"] = f"='{year}'!AE{21+i}" 
            ws_dash[f"Y{6+i}"] = f"=SUM('{year}'!AF{21+i}:AI{21+i})"
            
        # Chart
        chart = LineChart(); chart.title = texts['chart_title']; chart.style = 13
        chart.y_axis.title = texts['chart_y']; chart.y_axis.delete = False
        chart.x_axis.title = texts['chart_x']; chart.x_axis.delete = False
        chart.height = 12; chart.width = 26; chart.legend = None

        chart.d_table = DataTable(); chart.d_table.showHorzBorder = True
        chart.d_table.showVertBorder = True; chart.d_table.showOutline = True; chart.d_table.showKeys = True
        
        data = Reference(ws_dash, min_col=25, min_row=5, max_col=25, max_row=17)
        cats = Reference(ws_dash, min_col=24, min_row=6, max_col=24, max_row=17)
        chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
        ws_dash.add_chart(chart, "D6")

        # Info Center
        start_r = 30
        ws_dash.cell(row=start_r, column=4, value=texts['info_title']).font = Font(size=16, bold=True, color=c["navy"])
        for idx, (t, d) in enumerate(texts['manual']):
            r = start_r + 2 + (idx * 3)
            ct = ws_dash.cell(row=r, column=4, value=t); ct.font = Font(bold=True, color=c["blue"], size=11); ct.alignment = Alignment(vertical='top')
            ws_dash.merge_cells(start_row=r, start_column=5, end_row=r+1, end_column=15)
            cd = ws_dash.cell(row=r, column=5, value=d); cd.font = Font(color=c["gray"], size=10); cd.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_dash.column_dimensions["D"].width = 28
        for i in range(5,16): ws_dash.column_dimensions[get_column_letter(i)].width = 9

    wb.save(FILENAME)
    return FILENAME